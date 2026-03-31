import os
import openpyxl
from typing import List, Tuple, Set, Any
from langchain_core.documents import Document

class ExcelProcessor:
    """Excelファイル（.xlsx）をRAG向けに最適化して読み込むプロセッサ"""

    # ====================================================================
    # コントローラー（入口）
    # ====================================================================
    def process(self, file_path: str, original_filename: str = None) -> List[Document]:
        """
        ファイル名から判断して適切なプロセスメソッドにルーティングする入口。
        """
        filename = original_filename if original_filename else os.path.basename(file_path)
        
        if filename.startswith("TypeB"):
            print("(TypeB)のExcelを解析します")
            return self.process_typeB(
                file_path=file_path, 
                original_filename=original_filename,
                header_row=9,       # ヘッダーの開始行
                data_start_row=14,   # 実際のデータの開始行
                message_cols=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]
            )
        elif filename.startswith("TypeC"):
            print("(TypeC)のExcelを解析します")
            return self.process_typeC(file_path, original_filename)
        else:
            # ファイル名が "TypeA" で始まる場合、またはそれ以外の場合はデフォルトの汎用処理へ
            print("(TypeA)のExcelを解析します")
            return self.process_typeA(file_path, original_filename)
        
    # ====================================================================
    # 共通処理（ヘルパーメソッド）
    # ====================================================================

    def _should_skip_sheet(self, ws) -> bool:
        """処理をスキップすべきシートか判定する"""
        if ws.sheet_state == 'hidden':
            return True
        if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
            return True
        if ws.max_row <= 1:
            return True
        return False
    
    def _clean_value(self, val: Any) -> str:
        """値をクリーニング：改行を除去、連続空白を1つに"""
        if val is None:
            return ""
        val_str = str(val).replace('\r\n', '').replace('\n', '').replace('\r', '')
        return ' '.join(val_str.split()).strip()
    
    def _extract_sheet_matrix(self, ws) -> Tuple[List[List[Any]], int, int, Set[Tuple[int, int]]]:
        """
        セル結合を展開した2次元配列（マトリックス）と、縦結合のセル情報を返す
        """
        max_r = ws.max_row
        max_c = ws.max_column

        # マトリックス初期化
        matrix = [[None for _ in range(max_c)] for _ in range(max_r)]
        
        # ▼【追加箇所1】縦方向の結合セルを記憶するセットを準備
        vertical_merged_cells = set()
        
        # 値の格納
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                matrix[r-1][c-1] = ws.cell(row=r, column=c).value
                
        # セル結合の展開（結合範囲すべてに左上の値をコピー）
        for merged_range in ws.merged_cells.ranges:
            min_c, min_r, max_c_merge, max_r_merge = merged_range.bounds
            top_left_val = ws.cell(row=min_r, column=min_c).value
            
            # ▼【追加箇所2】縦方向のセル結合を記憶 (TypeCで使用)
            if max_r_merge > min_r:
                for r in range(min_r, max_r_merge + 1):
                    for c in range(min_c, max_c_merge + 1):
                        vertical_merged_cells.add((r-1, c-1))

            # 値の展開コピー
            for r in range(min_r, max_r_merge + 1):
                for c in range(min_c, max_c_merge + 1):
                    if 0 <= r-1 < max_r and 0 <= c-1 < max_c:
                        matrix[r-1][c-1] = top_left_val

        # ▼【追加箇所3】戻り値を4つにする（最後に vertical_merged_cells を追加）
        return matrix, max_r, max_c, vertical_merged_cells
    
    # ====================================================================
    # メソッド1：(TypeA)Excelの場合
    # ====================================================================
    def process_typeA(self, file_path: str, original_filename: str = None) -> List[Document]:
        documents = []
        filename = original_filename if original_filename else os.path.basename(file_path)
        
        # スキップ対象のシート名を定義（完全一致）
        SKIP_SHEETS = {"表紙", "変更履歴", "全体"}

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                
                # 0. 特定のシート名に一致した場合は処理をスキップ
                if sheet_name in SKIP_SHEETS:
                    continue
                
                ws = wb[sheet_name]
                if self._should_skip_sheet(ws): continue
                
                # 非表示行の特定
                hidden_rows = set()
                for r_idx, row_dim in ws.row_dimensions.items():
                    if row_dim.hidden:
                        hidden_rows.add(r_idx)
                
                matrix, max_r, max_c, _ = self._extract_sheet_matrix(ws)
                
                # 1. マトリックス構築とセル結合の展開
                matrix = [[None for _ in range(max_c)] for _ in range(max_r)]
                for r in range(1, max_r + 1):
                    for c in range(1, max_c + 1):
                        matrix[r-1][c-1] = ws.cell(row=r, column=c).value
                        
                for merged_range in ws.merged_cells.ranges:
                    min_c, min_r, max_c_merge, max_r_merge = merged_range.bounds
                    top_left_val = ws.cell(row=min_r, column=min_c).value
                    for r in range(min_r, max_r_merge + 1):
                        for c in range(min_c, max_c_merge + 1):
                            matrix[r-1][c-1] = top_left_val

                # 2. 表の開始位置（ヘッダー）の自動検知
                header_start_idx = -1
                MIN_UNIQUE_COLS_FOR_TABLE = 3 
                
                for i, row in enumerate(matrix):
                    r_excel = i + 1
                    if r_excel in hidden_rows: continue
                    
                    unique_vals = set(str(v).strip() for v in row if v is not None and str(v).strip() != "")
                    if len(unique_vals) >= MIN_UNIQUE_COLS_FOR_TABLE:
                        header_start_idx = i
                        break
                        
                if header_start_idx == -1:
                    continue # 表が見つからないシートはスキップ

                # 3. ヘッダーの階層検知
                header_end_idx = header_start_idx
                for merged_range in ws.merged_cells.ranges:
                    min_c, min_r, max_c_merge, max_r_merge = merged_range.bounds
                    if min_r - 1 <= header_start_idx <= max_r_merge - 1:
                        if (max_r_merge - 1) > header_end_idx:
                            header_end_idx = max_r_merge - 1

                # 4. メタデータ（説明文）抽出
                metadata_lines = []
                for i in range(header_start_idx):
                    r_excel = i + 1
                    if r_excel in hidden_rows: continue
                    
                    row_vals = []
                    for v in matrix[i]:
                        val_str = str(v).strip()
                        if v is not None and val_str != "" and val_str not in row_vals:
                            row_vals.append(val_str)
                    if row_vals:
                        metadata_lines.append(" ".join(row_vals))
                metadata_text = "\n".join(metadata_lines)

                # 5. ヘッダー平坦化
                flat_headers = []
                for c in range(max_c):
                    col_header_parts = []
                    for r in range(header_start_idx, header_end_idx + 1):
                        val = matrix[r][c]
                        if val is not None and str(val).strip() != "":
                            val_str = str(val).strip().replace('\n', '')
                            if not col_header_parts or col_header_parts[-1] != val_str:
                                col_header_parts.append(val_str)
                    flat_headers.append("_".join(col_header_parts) if col_header_parts else f"列{c+1}")

                # 6. データ行抽出とLangChain Document化
                for r in range(header_end_idx + 1, max_r):
                    r_excel = r + 1
                    if r_excel in hidden_rows: continue
                        
                    row_data = matrix[r]
                    unique_data_vals = set(str(v).strip() for v in row_data if v is not None and str(v).strip() != "")
                    if len(unique_data_vals) < 2:
                        continue # フィルタ用の空行などはスキップ
                        
                    lines = []
                    lines.append(f"■ファイル名: {filename}")
                    lines.append(f"■シート名: {sheet_name}")
                    if metadata_text:
                        lines.append(f"■シート前提情報:\n{metadata_text}\n")
                        
                    for c in range(max_c):
                        header = flat_headers[c]
                        val = row_data[c]
                        val_str = "(空欄)" if val is None or str(val).strip() == "" else str(val).strip().replace('\n', ' ')
                        lines.append(f"{header}: {val_str}")
                        
                    page_content = "\n".join(lines)
                    
                    # 1行のデータを1つのLangChain Documentとして作成
                    # メタデータにシート番号（page_numとして代用）を付与
                    doc = Document(
                        page_content=page_content,
                        metadata={
                            "source": file_path,
                            "page_num": sheet_idx + 1,
                            "sheet_name": sheet_name,
                            "total_pages": len(wb.sheetnames)
                        }
                    )
                    documents.append(doc)

        except Exception as e:
            # エラーロギング（既存のloggerがあればそれを使用してください）
            print(f"Excel処理中にエラーが発生しました: {e}")
            raise

        return documents
    
    # ====================================================================
    # メソッド2：(TypeB)gncanvehcs.xlsx用
    # ====================================================================
    def process_typeB(self, file_path: str, original_filename: str = None,
                header_row: int = None, data_start_row: int = None,
                message_cols: List[int] = None) -> List[Document]:
        documents = []
        filename = original_filename if original_filename else os.path.basename(file_path)

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]
                if ws.sheet_state == 'hidden': 
                    continue
                
                max_r = ws.max_row
                max_c = ws.max_column
                if max_r <= 1:
                    continue

                hidden_rows = set(r_idx for r_idx, dim in ws.row_dimensions.items() if dim.hidden)

                # ============================================
                # 1. マトリックス構築（セル結合展開）
                # ============================================
                raw_matrix = [[None for _ in range(max_c)] for _ in range(max_r)]
                for r in range(1, max_r + 1):
                    for c in range(1, max_c + 1):
                        raw_matrix[r-1][c-1] = ws.cell(row=r, column=c).value

                # セル結合を展開
                matrix = [row[:] for row in raw_matrix]
                for merged_range in ws.merged_cells.ranges:
                    min_c, min_r, max_c_m, max_r_m = merged_range.bounds
                    top_left_val = ws.cell(row=min_r, column=min_c).value
                    for r in range(min_r, max_r_m + 1):
                        for c in range(min_c, max_c_m + 1):
                            if 0 <= r-1 < max_r and 0 <= c-1 < max_c:
                                matrix[r-1][c-1] = top_left_val

                # ============================================
                # 2. ヘッダー・データ範囲の決定
                # ============================================
                if header_row is not None and data_start_row is not None:
                    header_start_idx = header_row - 1
                    header_end_idx = data_start_row - 2
                    data_start_idx = data_start_row - 1
                else:
                    header_start_idx, header_end_idx, data_start_idx = \
                        self._detect_range_auto_v2(matrix, hidden_rows, max_r, max_c)

                if header_start_idx < 0:
                    continue

                # ============================================
                # 3. メタデータ抽出
                # ============================================
                metadata_lines = []
                for i in range(header_start_idx):
                    if (i + 1) in hidden_rows:
                        continue
                    row_text = " ".join(
                        self._clean_value(v) for v in matrix[i] 
                        if v is not None and str(v).strip() != ""
                    )
                    if row_text:
                        metadata_lines.append(row_text)
                metadata_text = "\n".join(metadata_lines)

                # ============================================
                # 4. 親階層を意識した横方向補完と平坦化（★修正箇所）
                # ============================================
                
                header_raw = []
                for r in range(header_start_idx, header_end_idx + 1):
                    row = []
                    for c in range(max_c):
                        val = matrix[r][c] if c < len(matrix[r]) else None
                        val_str = self._clean_value(val) if val is not None else None
                        row.append(val_str if val_str != "" else None)
                    header_raw.append(row)

                num_header_rows = len(header_raw)
                
                # 4.1 安全な横方向補完（左からの継承）
                for r in range(num_header_rows):
                    # 【重要】一番下の行（列の固有項目名）は横に継承しない（侵食防止）
                    if r == num_header_rows - 1:
                        continue
                        
                    current_val = None
                    for c in range(max_c):
                        if header_raw[r][c] is not None:
                            current_val = header_raw[r][c]
                        else:
                            # 上の階層（親グループ）が変わった場合は、継承をリセットする
                            can_inherit = True
                            if r > 0 and c > 0:
                                parent_curr = header_raw[r-1][c]
                                parent_prev = header_raw[r-1][c-1]
                                if parent_curr != parent_prev:
                                    current_val = None
                                    can_inherit = False
                            
                            if can_inherit and current_val is not None:
                                header_raw[r][c] = current_val

                # 4.2 縦方向に結合して平坦化
                flat_headers = []
                for c in range(max_c):
                    parts = []
                    for r in range(num_header_rows):
                        val = header_raw[r][c]
                        if val is None:
                            continue
                        
                        # 連続する同じ名前の階層はスキップ
                        if parts and parts[-1] == val:
                            continue
                        
                        parts.append(val)
                    
                    if parts:
                        flat_headers.append("_".join(parts))
                    else:
                        flat_headers.append(f"列{c+1}")

                # ============================================
                # 5. Message Inf. 列の特定（伝搬対象）
                # ============================================
                if message_cols is None:
                    message_cols = self._detect_message_columns(flat_headers)

                # ============================================
                # 6. データ行の処理（伝搬ロジック）
                # ============================================
                num_cols = min(max_c, len(flat_headers))
                data_rows = []
                current_message_values = {}
                
                for r in range(data_start_idx, max_r):
                    if (r + 1) in hidden_rows:
                        continue
                    
                    row = matrix[r][:num_cols] if len(matrix[r]) >= num_cols else matrix[r] + [None] * (num_cols - len(matrix[r]))
                    
                    # 空行チェック
                    if all(v is None or str(v).strip() == "" for v in row):
                        continue
                    
                    # 新規メッセージ判定
                    is_new_message = self._is_new_message_row(row, message_cols, matrix, r)
                    
                    if is_new_message:
                        current_message_values = {}
                        for c in message_cols:
                            if c < len(row):
                                val = row[c]
                                if val is not None and str(val).strip() != "":
                                    current_message_values[c] = val
                    else:
                        # 伝搬適用
                        for c in message_cols:
                            if c < len(row):
                                current_val = row[c]
                                if (current_val is None or str(current_val).strip() == "") and c in current_message_values:
                                    row[c] = current_message_values[c]
                                elif current_val is not None and str(current_val).strip() != "":
                                    current_message_values[c] = current_val
                    
                    data_rows.append((r, row, is_new_message))

                # ============================================
                # 7. Document生成
                # ============================================
                for excel_row_num, row, is_new_msg in data_rows:
                    lines = []
                    lines.append(f"■ファイル名: {filename}")
                    lines.append(f"■シート名: {sheet_name}")
                    if metadata_text:
                        lines.append(f"■シート前提情報:\n{metadata_text}")
                    
                    row_entries = []
                    for c in range(min(len(row), len(flat_headers))):
                        val = row[c]
                        if val is None:
                            continue
                        
                        val_str = self._clean_value(val)
                        if val_str == "":
                            continue
                        
                        header = flat_headers[c]
                        row_entries.append(f"{header}: {val_str}")
                    
                    if row_entries:
                        lines.append("")
                        lines.extend(row_entries)
                    
                    doc = Document(
                        page_content="\n".join(lines),
                        metadata={
                            "source": filename,
                            "sheet_name": sheet_name,
                            "row_number": excel_row_num + 1,
                            "is_message_header": is_new_msg
                        }
                    )
                    documents.append(doc)

        except Exception as e:
            print(f"Excel処理エラー: {e}")
            import traceback
            traceback.print_exc()
            raise

        return documents

    def _clean_value(self, val) -> str:
        """値をクリーニング：改行を除去、連続空白を1つに"""
        if val is None:
            return ""
        val_str = str(val)
        # 縦書き(B\nA\nT)に対応するため、改行はスペースではなく除去する
        val_str = val_str.replace('\r\n', '').replace('\n', '').replace('\r', '')
        # 連続スペースを1つに（元からあったスペースは残る）
        val_str = ' '.join(val_str.split())
        return val_str.strip()

    def _detect_range_auto_v2(self, matrix, hidden_rows, max_r, max_c):
        """ヘッダー・データ範囲検出（前回と同じ実装を維持）"""
        header_start = -1
        header_end = -1
        data_start = -1
        
        for i in range(min(20, max_r)):
            if (i + 1) in hidden_rows: continue
            row_text = " ".join([str(v).upper() for v in matrix[i] if v is not None])
            if 'MSG' in row_text and 'LABEL' in row_text:
                header_start = i
                break
        
        if header_start >= 0:
            for i in range(header_start + 1, min(header_start + 10, max_r)):
                if (i + 1) in hidden_rows: continue
                first_val = matrix[i][0] if len(matrix[i]) > 0 else None
                if first_val and str(first_val).strip() != "" and len(str(first_val)) >= 3:
                    data_start = i
                    header_end = i - 1
                    break
                    
        if header_start < 0:
            return 9, 13, 14
        if data_start < 0:
            header_end = header_start + 1
            data_start = header_end + 1
            
        return header_start, header_end, data_start

    def _detect_message_columns(self, flat_headers):
        message_cols = []
        for c, h in enumerate(flat_headers):
            h_upper = h.upper().replace(' ', '').replace('_', '')
            score = 0
            if 'MESSAGEINF' in h_upper: score += 20
            elif 'MSG' in h_upper:
                if any(x in h_upper for x in ['LABEL', 'VARI', 'NAME', 'CANID', 'DLC']): score += 15
                elif any(x in h_upper for x in ['FDF', 'COM', 'APP', 'CYCLE', 'EVENT']): score += 10
            if 'TRANS' in h_upper and 'DATA' not in h_upper: score += 5
            if 'CHECKSUM' in h_upper or 'KZK' in h_upper: score += 10
            if score >= 5: message_cols.append(c)
        return sorted(set(message_cols))

    def _is_new_message_row(self, row, message_cols, matrix, row_idx):
        if not message_cols: return False
        first_c = message_cols[0]
        if first_c >= len(row): return False
        val = row[first_c]
        if val is None or str(val).strip() == "": return False
        if row_idx <= 0: return True
        prev_val = matrix[row_idx-1][first_c] if first_c < len(matrix[row_idx-1]) else None
        curr_str = str(val).strip()
        prev_str = str(prev_val).strip() if prev_val is not None else ""
        return curr_str != prev_str and curr_str != ""
    
    
    # ====================================================================
    # メソッド3：(TypeC)分析結果.xlsx用
    # ====================================================================
    def process_typeC(self, file_path: str, original_filename: str = None) -> List[Document]:
        """
        1つのシート内に複数の表（ブロック）が縦に点在しているExcelを処理します。
        最上位の表タイトルをすべての項目にプレフィックスとして付与し、完全な階層構造を作ります。
        """
        documents = []
        filename = original_filename if original_filename else os.path.basename(file_path)

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]
                if ws.sheet_state == 'hidden': continue
                max_r = ws.max_row
                max_c = ws.max_column
                if max_r == 1 and max_c == 1 and ws.cell(1,1).value is None: continue
                hidden_rows = set(r_idx for r_idx, dim in ws.row_dimensions.items() if dim.hidden)

                # 縦方向のセル結合を記憶
                vertical_merged_cells = set()
                for merged_range in ws.merged_cells.ranges:
                    min_c, min_r, max_c_merge, max_r_merge = merged_range.bounds
                    if max_r_merge > min_r:
                        for r in range(min_r, max_r_merge + 1):
                            for c in range(min_c, max_c_merge + 1):
                                vertical_merged_cells.add((r-1, c-1))

                # マトリックス構築とセル結合の展開（ここで「反映部品」などは右の列まで正しく展開されます）
                matrix = [[None for _ in range(max_c)] for _ in range(max_r)]
                for r in range(1, max_r + 1):
                    for c in range(1, max_c + 1):
                        matrix[r-1][c-1] = ws.cell(row=r, column=c).value
                        
                for merged_range in ws.merged_cells.ranges:
                    min_c, min_r, max_c_merge, max_r_merge = merged_range.bounds
                    top_left_val = ws.cell(row=min_r, column=min_c).value
                    for r in range(min_r, max_r_merge + 1):
                        for c in range(min_c, max_c_merge + 1):
                            matrix[r-1][c-1] = top_left_val

                # 空白行で表をブロック分割
                blocks = []
                current_block = []
                for r in range(max_r):
                    if (r + 1) in hidden_rows: continue
                    row = matrix[r]
                    valid_vals = [v for v in row if v is not None and str(v).strip() not in ("", "-", "_")]
                    if len(valid_vals) > 0:
                        current_block.append((r, row))
                    else:
                        if current_block:
                            blocks.append(current_block)
                            current_block = []
                if current_block: blocks.append(current_block)

                # 各表ブロックの処理
                for block in blocks:
                    if len(block) < 2: continue

                    # A. 表タイトルの抽出
                    table_title = ""
                    for v in block[0][1]:
                        if v is not None and str(v).strip() != "":
                            table_title = str(v).strip()
                            break
                            
                    header_start_idx = 0

                    # B. ヘッダー行数の自動判定
                    header_end_idx = header_start_idx
                    for i in range(header_start_idx + 1, len(block)):
                        prev_r_idx = block[i-1][0]
                        curr_r_idx = block[i][0]
                        prev_row = block[i-1][1]
                        curr_row = block[i][1]
                        
                        is_vertical_merge = any(
                            (prev_r_idx, c) in vertical_merged_cells and (curr_r_idx, c) in vertical_merged_cells
                            for c in range(max_c)
                        )
                        
                        prev_unique = len(set(str(v).strip() for v in prev_row if v is not None and str(v).strip() != ""))
                        curr_unique = len(set(str(v).strip() for v in curr_row if v is not None and str(v).strip() != ""))
                        is_expansion = (prev_unique < curr_unique) and (prev_unique <= 10)
                        
                        if is_vertical_merge or is_expansion:
                            header_end_idx = i
                        else:
                            break
                            
                    if header_end_idx + 1 >= len(block): continue

                    # ==========================================
                    # 【削除】強引な「C. 横方向の補完」ロジックは撤廃しました
                    # Excel本来のセル結合情報だけを信用します。
                    # ==========================================

                    # D. 階層ヘッダーの平坦化（ルート付与＆重複排除）
                    flat_headers = []
                    for c in range(max_c):
                        parts = []
                        # 常に一番親の要素（表タイトル）をセット
                        if table_title:
                            parts.append(table_title)

                        for i in range(header_start_idx, header_end_idx + 1):
                            val = matrix[block[i][0]][c]
                            val_str = str(val).strip().replace('\n', '') if val is not None else ""
                            
                            # 空欄でなく、直前の文字と同じでない場合のみ追加
                            if val_str != "" and (not parts or parts[-1] != val_str):
                                parts.append(val_str)
                                
                        flat_headers.append("_".join(parts) if parts else f"列{c+1}")

                    # E. データ行のチャンク化
                    for i in range(header_end_idx + 1, len(block)):
                        r_idx, row_data = block[i]
                        unique_data = set(str(v).strip() for v in row_data if v is not None and str(v).strip() != "")
                        if len(unique_data) < 2: continue

                        lines = [f"■ファイル名: {filename}", f"■シート名: {sheet_name}"]
                        lines.append("-" * 20)
                            
                        for c in range(max_c):
                            if row_data[c] is None or str(row_data[c]).strip() == "": continue
                            
                            header = flat_headers[c]
                            val_str = str(row_data[c]).strip().replace('\n', ' ')
                            lines.append(f"{header}: {val_str}")
                            
                        documents.append(Document(
                            page_content="\n".join(lines),
                            metadata={"source": filename, "sheet_name": sheet_name}
                        ))
        except Exception as e:
            print(f"Excel処理中にエラーが発生しました: {e}")
            raise
        return documents